from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
import os
import openpyxl
import re
import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
from reportlab.lib import colors
from reportlab.lib.units import inch
import requests

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.permanent_session_lifetime = timedelta(minutes=30)

# Configuración de la sesión (¡nuevo!)
app.config['SESSION_PERMANENT'] = True 

# Ruta predeterminada del archivo Excel
archivo_excel_path = os.path.join(os.getcwd(), "base de datos.xlsx")

@app.route('/')
def index():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    usuario = request.form['usuario']
    contraseña = request.form['contraseña']
    resultado_verificacion = verificar_usuario(usuario, contraseña)

    if resultado_verificacion:
        session.permanent = True
        session['logged_in'] = True  # Asegúrate de que esta línea esté presente
        session['nombre_usuario'] = usuario
        session['result'] = []
        api_key = obtener_api_key(usuario)
        session['api_key'] = api_key
        session.modified = True  # Guardar cambios en la sesión permanente

        if not api_key:
            return redirect(url_for('set_api'))
        else:
            flash('API cargada correctamente.', 'success')
            return redirect(url_for('analyze'))
    else:
        flash('Usuario o contraseña incorrectos.')
        return redirect(url_for('index'))


@app.route('/logout')
def logout():
    session.clear()  # Asegúrate de limpiar toda la sesión
    return redirect(url_for('index'))

@app.route('/analyze', methods=['GET', 'POST'])
def analyze():
    if 'nombre_usuario' not in session:
        return redirect(url_for('index'))

    if 'api_key' not in session or not session['api_key']:
        flash('Debe configurar su API Key antes de continuar.')
        return redirect(url_for('set_api'))

    if request.method == 'POST':
        xml_content = request.form['xml_content']
        if not (xml_content.startswith('<AppHdr xmlns="urn:iso:std:iso:20022:tech:xsd:head.001.001.02">') or xml_content.strip().endswith('</Document>')):
            flash('CONTENIDO NO ES XML')
            return redirect(url_for('analyze'))

        country_codes = obtener_codigos_paises()
        referencias_del_giro = []  # Inicializa aquí
        analysis_result = analyze_xml(xml_content, country_codes, referencias_del_giro)
        session['result'].extend(analysis_result)
        return render_template('report.html', analysis_result=analysis_result)

    return render_template('analyze.html')

@app.route('/new_analysis', methods=['GET'])
def new_analysis():
    if 'nombre_usuario' not in session:
        return redirect(url_for('index'))
    return render_template('analyze.html')

@app.route('/profile')
def profile():
    if 'nombre_usuario' not in session:
        return redirect(url_for('index'))
    api_key = session.get('api_key', '')
    return render_template('profile.html', api_key=api_key)

@app.route('/update_password', methods=['POST'])
def update_password():
    if 'nombre_usuario' not in session:
        return redirect(url_for('index'))
    
    new_password = request.form.get('new_password')
    usuario = session['nombre_usuario']
    
    if new_password:
        actualizar_contraseña(usuario, new_password)
        flash('Contraseña actualizada correctamente.', 'success')

    return redirect(url_for('profile'))

@app.route('/update_api', methods=['POST'])
def update_api():
    if 'nombre_usuario' not in session:
        return redirect(url_for('index'))
    
    api_key = request.form.get('api_key')
    usuario = session['nombre_usuario']
    
    if api_key:
        guardar_api_key(usuario, api_key)
        session['api_key'] = api_key
        flash('API Key actualizada correctamente.', 'success')

    return redirect(url_for('profile'))

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        usuario = request.form['usuario']
        usuarios = obtener_usuarios()
        if any(user['Usuario'].lower() == usuario.lower() for user in usuarios):
            session['reset_usuario'] = usuario
            return redirect(url_for('reset_password'))
        else:
            flash('Usuario no encontrado.')
            return redirect(url_for('forgot_password'))
    return render_template('forgot_password.html')

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    if request.method == 'POST':
        nueva_contraseña = request.form['new_password']
        usuario = session.get('reset_usuario')
        if usuario and nueva_contraseña:
            actualizar_contraseña(usuario, nueva_contraseña)
            session.pop('reset_usuario', None)
            flash('Contraseña restablecida correctamente.', 'success')
            return redirect(url_for('index'))
        else:
            flash('Error al restablecer la contraseña.', 'danger')
            return redirect(url_for('reset_password'))
    return render_template('reset_password.html')


def actualizar_contraseña(usuario, new_password):
    try:
        workbook = openpyxl.load_workbook(archivo_excel_path)
        sheet = workbook["usuarios"]
        for row in sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value.lower() == usuario.lower():
                row[1].value = new_password
                break
        workbook.save(archivo_excel_path)
    except Exception as e:
        print(f"Error: {e}")

@app.after_request
def add_header(response):
    response.cache_control.no_store = True
    return response



@app.route('/daily_report', methods=['POST'])
def daily_report():
    if 'api_key' not in session or not session['api_key']:
        flash('Debe configurar su API Key antes de continuar.')
        return redirect(url_for('set_api'))

    format_type = request.form['format']
    result = session.get('result', [])
    report_values = prepare_report_data(result)
    headers = ["Referencia del Giro", "Banco Beneficiario", "Nombre del Beneficiario", "Monto de la Transacción", "Motivo", "Tipo de Giro", "País Beneficiario"]

    if format_type == 'excel':
        output = BytesIO()
        df = pd.DataFrame(report_values, columns=headers)
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Reporte', startrow=1, header=False)
            workbook = writer.book
            worksheet = writer.sheets['Reporte']
            cell_format = workbook.add_format({'border': 1})
            for row_num, row in enumerate(df.values, 1):
                for col_num, value in enumerate(row):
                    worksheet.write(row_num, col_num, value, cell_format)
            header_format = workbook.add_format({'bold': True, 'text_wrap': True, 'valign': 'top', 'fg_color': '#D7E4BC', 'border': 1})
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
            for column in df:
                column_width = max(df[column].astype(str).apply(len).max(), len(column))
                col_idx = df.columns.get_loc(column)
                worksheet.set_column(col_idx, col_idx, column_width + 1)
            worksheet.set_footer(f'&LUsuario: {session["nombre_usuario"]}&C&F')
        output.seek(0)
        return send_file(output, download_name="report.xlsx", as_attachment=True)
    elif format_type == 'pdf':
        output = BytesIO()
        generate_pdf(output, report_values)
        output.seek(0)
        return send_file(output, download_name="Reporte_Diario_RCSA.pdf", as_attachment=True)

@app.route('/set_api', methods=['GET', 'POST'])
def set_api():
    if 'nombre_usuario' not in session:
        return redirect(url_for('index'))

    if request.method == 'POST':
        api_key = request.form['api_key']
        usuario = session['nombre_usuario']
        guardar_api_key(usuario, api_key)
        session['api_key'] = api_key
        flash('API Key guardada correctamente.', 'success')
        return redirect(url_for('analyze'))

    return render_template('set_api.html')

@app.route('/generate_report', methods=['POST'])
def generate_report():
    if 'api_key' not in session or not session['api_key']:
        flash('Debe configurar su API Key antes de continuar.')
        return redirect(url_for('set_api'))

    format_type = request.form['format']
    result = session.get('result', [])
    report_values = prepare_report_data(result)
    headers = ["Referencia del Giro", "Banco Beneficiario", "Nombre del Beneficiario", "Monto de la Transacción", "Motivo", "Tipo de Giro", "País Beneficiario"]

    if format_type == 'excel':
        output = BytesIO()
        df = pd.DataFrame(report_values, columns=headers)
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Reporte', startrow=1, header=False)
            workbook = writer.book
            worksheet = writer.sheets['Reporte']
            cell_format = workbook.add_format({'border': 1})
            for row_num, row in enumerate(df.values, 1):
                for col_num, value in enumerate(row):
                    worksheet.write(row_num, col_num, value, cell_format)
            header_format = workbook.add_format({'bold': True, 'text_wrap': True, 'valign': 'top', 'fg_color': '#D7E4BC', 'border': 1})
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
            for column in df:
                column_width = max(df[column].astype(str).apply(len).max(), len(column))
                col_idx = df.columns.get_loc(column)
                worksheet.set_column(col_idx, col_idx, column_width + 1)
            worksheet.set_footer(f'&LUsuario: {session["nombre_usuario"]}&C&F')
        output.seek(0)
        return send_file(output, download_name="report.xlsx", as_attachment=True)
    elif format_type == 'pdf':
        output = BytesIO()
        generate_pdf(output, report_values)
        output.seek(0)
        return send_file(output, download_name="ReporteXML.pdf", as_attachment=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xml'}

def verificar_usuario(usuario, contraseña):
    usuarios = obtener_usuarios()
    for user in usuarios:
        if user['Usuario'].lower() == usuario.lower() and user['Contraseña'] == contraseña:
            return True
    return False

def obtener_usuarios():
    try:
        workbook = openpyxl.load_workbook(archivo_excel_path)
        sheet = workbook["usuarios"]
        usuarios = [{'Usuario': str(row[0]), 'Contraseña': str(row[1]), 'API': row[2] if len(row) > 2 else None} for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] and row[1]]
        return usuarios
    except Exception as e:
        print(f"Error: {e}")
        return []

def obtener_api_key(usuario):
    usuarios = obtener_usuarios()
    for user in usuarios:
        if user['Usuario'].lower() == usuario.lower():
            return user.get('API')
    return None

def guardar_api_key(usuario, api_key):
    try:
        workbook = openpyxl.load_workbook(archivo_excel_path)
        sheet = workbook["usuarios"]
        for row in sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value.lower() == usuario.lower():
                row[2].value = api_key
                break
        workbook.save(archivo_excel_path)
    except Exception as e:
        print(f"Error: {e}")

def cambiar_contraseña(usuario, nueva_contraseña):
    try:
        workbook = openpyxl.load_workbook(archivo_excel_path)
        sheet = workbook["usuarios"]
        for row in sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value.lower() == usuario.lower():
                row[1].value = nueva_contraseña
                break
        workbook.save(archivo_excel_path)
    except Exception as e:
        print(f"Error: {e}")

def obtener_codigos_paises():
    try:
        workbook = openpyxl.load_workbook(archivo_excel_path)
        sheet = workbook["paises"]
        country_codes = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            country_codes[row[0]] = row[1]
        return country_codes
    except Exception as e:
        print(f"Error: {e}")
        return {}

def obtener_informacion_banco(swift_code):
    swift_code = swift_code.rstrip('X')  # Elimina los 'X' del final si existen
    api_url = f'https://api.api-ninjas.com/v1/swiftcode?swift={swift_code}'
    response = requests.get(api_url, headers={'X-Api-Key': session.get('api_key')})
    if response.status_code == 200:
        bank_info = response.json()
        if bank_info:
            return bank_info[0]['bank_name']
    return "N/A"

def analyze_xml(xml_content, country_codes, referencias_del_giro):
    result = session.get('result', [])
    xml_content = preprocess_xml(xml_content)
    wrapped_xml_content = f"<root>{xml_content}</root>"
    try:
        root = ET.ElementTree(ET.fromstring(wrapped_xml_content)).getroot()
        ns = {
            'head': 'urn:iso:std:iso:20022:tech:xsd:head.001.001.02',
            'pacs': 'urn:iso:std:iso:20022:tech:xsd:pacs.008.001.08'
        }
        bicfi_from = flexible_find(root, './/head:Fr/head:FIId/head:FinInstnId/head:BICFI', ns)
        bicfi_to = flexible_find(root, './/head:To/head:FIId/head:FinInstnId/head:BICFI', ns)
        bicfi_beneficiary = flexible_find(root, './/pacs:CdtrAgt/pacs:FinInstnId/pacs:BICFI', ns)
        biz_msg_idr = flexible_find(root, './/head:BizMsgIdr', ns)
        debtor_account_number = root.find('.//pacs:DbtrAcct/pacs:Id/pacs:Othr/pacs:Id', ns).text if root.find('.//pacs:DbtrAcct/pacs:Id/pacs:Othr/pacs:Id', ns) is not None else "N/A"
        amount_element = root.find('.//pacs:IntrBkSttlmAmt', ns)
        beneficiary_account_number = root.find('.//pacs:CdtrAcct/pacs:Id/pacs:Othr/pacs:Id', ns)
        beneficiary_account_number = beneficiary_account_number.text if beneficiary_account_number is not None else "N/A"
        amount = f"{amount_element.text} {amount_element.attrib['Ccy']}" if amount_element is not None else "N/A"
        debtor_name = flexible_find(root, './/pacs:Dbtr/pacs:Nm', ns)
        creditor_name = flexible_find(root, './/pacs:Cdtr/pacs:Nm', ns)
        debtor_address_element = flexible_find(root, './/pacs:Dbtr/pacs:PstlAdr', ns)
        creditor_address_element = flexible_find(root, './/pacs:Cdtr/pacs:PstlAdr', ns)

        if biz_msg_idr and biz_msg_idr[0].text not in referencias_del_giro:
            referencias_del_giro.append(biz_msg_idr[0].text)

        if creditor_address_element and len(creditor_address_element) > 0:
            creditor_address_details = extract_address_details(creditor_address_element[0], ns, country_codes)
            creditor_street = creditor_address_details["street"]
            creditor_city = creditor_address_details["city"]
            creditor_country = creditor_address_details["country"]
            ofac_present = creditor_address_details["ofac"]
            gafi_present = creditor_address_details["gafi"]
            paraiso_present = creditor_address_details["P.F"]

        if creditor_country in country_codes:
            pais_original = country_codes[creditor_country]
            creditor_country, ofac_present, gafi_present, paraiso_present = limpiar_nombre_pais(pais_original)

        sanciones = "No Mantiene"
        if ofac_present:
            sanciones = "OFAC, Revisar RCSA"
        elif gafi_present:
            sanciones = "GAFI, Revisar RCSA"
        elif paraiso_present:
            sanciones = "Paraiso Fiscal, Revisar"

        motive = flexible_find(root, './/pacs:RmtInf/pacs:Ustrd', ns)

        debtor_street = debtor_address_element[0].find('.//pacs:StrtNm', ns).text if debtor_address_element and debtor_address_element[0].find('.//pacs:StrtNm', ns) is not None else (debtor_address_element[0].find('.//pacs:AdrLine', ns)[0].text if debtor_address_element and debtor_address_element[0].find('.//pacs:AdrLine', ns) else "N/A")
        creditor_street = creditor_address_element[0].find('.//pacs:StrtNm', ns).text if creditor_address_element and creditor_address_element[0].find('.//pacs:StrtNm', ns) is not None else (creditor_address_element[0].find('.//pacs:AdrLine', ns)[0].text if creditor_address_element and creditor_address_element[0].find('.//pacs:AdrLine', ns) else "N/A")

        debtor_address_details = extract_address_details(debtor_address_element[0], ns, country_codes) if debtor_address_element else {"street": "N/A"}
        creditor_address_details = extract_address_details(creditor_address_element[0], ns, country_codes) if creditor_address_element else {"street": "N/A"}

        bank_from = obtener_informacion_banco(bicfi_from[0].text if bicfi_from else "N/A")
        bank_to = obtener_informacion_banco(bicfi_to[0].text if bicfi_to else "N/A")
        bank_beneficiary = obtener_informacion_banco(bicfi_beneficiary[0].text if bicfi_beneficiary else "No encontrado")

        country_code_from = bicfi_from[0].text[4:6] if bicfi_from else ""
        country_code_to = bicfi_to[0].text[4:6] if bicfi_to else ""
        country_code_beneficiary = bicfi_beneficiary[0].text[4:6] if bicfi_beneficiary else ""

        country_from = limpiar_nombre_pais_solo(country_codes.get(country_code_from, "Desconocido"))
        country_to = limpiar_nombre_pais_solo(country_codes.get(country_code_to, "Desconocido"))
        country_beneficiary = limpiar_nombre_pais_solo(country_codes.get(country_code_beneficiary, "Desconocido"))

        new_result = [
            ("Banco Emisor: ", bank_from, country_from),
            ("Banco Corresponsal: ", bank_to, country_to),
            ("Banco Beneficiario: ", bank_beneficiary, country_beneficiary),
            ("Referencia del giro: ", biz_msg_idr[0].text if biz_msg_idr else "N/A", ""),
            ("Nombre del Ordenante: ", debtor_name[0].text if debtor_name else "N/A", ""),
            ("Monto de la transacción: ", amount, ""),
            ("Número de Cuenta del Ordenante: ", debtor_account_number, ""),
            ("Dirección del Ordenante: ", debtor_address_details["street"], ""),
            ("Número de Cuenta del Beneficiario: ", beneficiary_account_number, ""),
            ("Nombre Del Beneficiario: ", creditor_name[0].text if creditor_name else "N/A", ""),
            ("Dirección del Beneficiario: ", creditor_address_details["street"], ""),
            ("Ciudad del Beneficiario: ", creditor_city, ""),
            ("País del Beneficiario: ", creditor_country, ""),
            ("Sancion País: ", sanciones, ""),
            ("Motivo: ", motive[0].text if motive else "N/A", ""),
            ("Tipo de Giro: ", determinar_tipo_giro(country_from), "")
        ]

        result.extend(new_result)
        session['result'] = result
        return new_result
    except ET.ParseError as e:
        return [("Error al procesar el XML: ", str(e), "")]

def preprocess_xml(xml_content):
    return re.sub(r'>\s+<', '><', xml_content)

def flexible_find(element, path, namespaces):
    try:
        return element.findall(path, namespaces)
    except SyntaxError:
        no_ns_path = re.sub(r'\w+:', '', path)
        return element.findall(no_ns_path)

def extract_address_details(address_element, namespaces, country_codes):
    street_text = city_text = country_text = "N/A"
    ofac_present = gafi_present = paraiso_present = False

    street = address_element.find('.//pacs:StrtNm', namespaces)
    city = address_element.find('.//pacs:TwnNm', namespaces)
    country = address_element.find('.//pacs:Ctry', namespaces)

    if street is not None or city is not None or country is not None:
        street_text = street.text if street is not None else "N/A"
        city_text = city.text if city is not None else "N/A"
        if country is not None and country.text in country_codes:
            country_text, ofac_present, gafi_present, paraiso_present = limpiar_nombre_pais(country_codes[country.text])
    else:
        address_lines = address_element.findall('.//pacs:AdrLine', namespaces)
        if address_lines:
            full_address = ", ".join(line.text for line in address_lines)
            street_text = full_address
            if " " in address_lines[-1].text:
                city_text, country_code = address_lines[-1].text.rsplit(" ", 1)
                if country_code.upper() in country_codes:
                    country_text, ofac_present, gafi_present, paraiso_present = limpiar_nombre_pais(country_codes[country_code.upper()])

    return {
        "street": street_text,
        "city": city_text,
        "country": country_text,
        "ofac": ofac_present,
        "gafi": gafi_present,
        "P.F": paraiso_present
    }

def limpiar_nombre_pais(nombre_pais):
    ofac_present = "(OFAC)" in nombre_pais
    gafi_present = "(GAFI)" in nombre_pais
    paraiso_present = "(P.F)" in nombre_pais
    nombre_pais_limpio = re.sub(r"\s*\([^)]*\)", "", nombre_pais).strip()
    return nombre_pais_limpio, ofac_present, gafi_present, paraiso_present

def limpiar_nombre_pais_solo(nombre_pais):
    return re.sub(r"\s*\([^)]*\)", "", nombre_pais).strip()

def determinar_tipo_giro(pais_banco_emisor):
    ecuador = "ECUADOR"
    if pais_banco_emisor.upper() == ecuador.upper():
        return "INPUT"
    else:
        return "OUTPUT"

def create_table_style(font_size):
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), font_size),
        ('FONTSIZE', (0, 1), (-1, -1), font_size),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ])

def create_total_table_style(font_size):
    return TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.beige),
        ('BACKGROUND', (1, 0), (1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), font_size),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])

def generate_pdf(output, report_values):
    report_values_upper = [[str(cell).upper() for cell in row] for row in report_values]
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    main_table_style = create_table_style(font_size=7)
    total_table_style = create_total_table_style(font_size=7)
    headers = ["REFERENCIA DEL GIRO", "BANCO BENEFICIARIO", "NOMBRE DEL BENEFICIARIO", "MONTO DE LA TRANSACCIÓN", "MOTIVO", "TIPO DE GIRO", "PAÍS BENEFICIARIO"]
    data = [headers] + report_values_upper
    try:
        total_amount = sum(float(row[3].split()[0]) for row in report_values)
    except Exception as e:
        total_amount = 0.0
    table = Table(data, repeatRows=1)
    table.setStyle(main_table_style)
    total_data = [["TOTAL", f"{total_amount:.2f} USD"]]
    total_table = Table(total_data, colWidths=[1.0 * inch, 2.0 * inch])
    total_table.setStyle(total_table_style)

    def header(canvas, doc):
        canvas.saveState()
        canvas.setFont('Times-Bold', 12)
        canvas.drawCentredString(doc.width / 2.0 + doc.leftMargin, doc.height + doc.topMargin - 10, "INFORME DIARIO")
        canvas.setFont('Times-Roman', 9)
        canvas.drawString(doc.leftMargin, doc.height + doc.topMargin - 20, f"USUARIO: {session['nombre_usuario'].upper()}")
        canvas.drawString(doc.width + doc.leftMargin - 100, doc.height + doc.topMargin - 20, "FECHA: " + datetime.now().strftime("%d-%m-%Y %H:%M").upper())
        canvas.restoreState()

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Times-Roman', 7)
        canvas.drawString(doc.leftMargin, 0.75 * inch, "PÁGINA %d" % doc.page)
        canvas.restoreState()

    elements = []
    elements.append(Spacer(1, 0.3 * inch))
    elements.append(table)
    elements.append(Spacer(1, 0.1 * inch))
    elements.append(total_table)
    doc.build(elements, onFirstPage=header, onLaterPages=footer)

def prepare_report_data(result):
    report_data = {}

    for item in result:
        label, value, _ = item

        if label.startswith("Referencia del giro:"):
            if value not in report_data:
                report_data[value] = {
                    "Referencia del Giro": value,
                    "Banco Beneficiario": None,
                    "Nombre del Beneficiario": None,
                    "Monto de la Transacción": None,
                    "Motivo": None,
                    "País Beneficiario": None,
                    "Tipo de Giro": None
                }
        elif label.startswith("Banco Beneficiario:"):
            for ref_giro, data in report_data.items():
                if data["Banco Beneficiario"] is None:
                    data["Banco Beneficiario"] = value
                    break
        elif label.startswith("Nombre Del Beneficiario:"):
            for ref_giro, data in report_data.items():
                if data["Nombre del Beneficiario"] is None:
                    data["Nombre del Beneficiario"] = value
                    break
        elif label.startswith("Monto de la transacción:"):
            for ref_giro, data in report_data.items():
                if data["Monto de la Transacción"] is None:
                    data["Monto de la Transacción"] = value
                    break
        elif label.startswith("Motivo:"):
            for ref_giro, data in report_data.items():
                if data["Motivo"] is None:
                    data["Motivo"] = value
                    break
        elif label.startswith("Tipo de Giro:"):
            for ref_giro, data in report_data.items():
                if data["Tipo de Giro"] is None:
                    data["Tipo de Giro"] = value
                    break
        elif label.startswith("País del Beneficiario:"):
            for ref_giro, data in report_data.items():
                if data["País Beneficiario"] is None:
                    data["País Beneficiario"] = value
                    break

    report_values = [
        [
            data["Referencia del Giro"],
            data["Banco Beneficiario"],
            data["Nombre del Beneficiario"],
            data["Monto de la Transacción"],
            data["Motivo"],
            data["Tipo de Giro"],
            data["País Beneficiario"]
        ]
        for _, data in report_data.items()
    ]

    return report_values

if __name__ == '__main__':
    app.run(debug=True)